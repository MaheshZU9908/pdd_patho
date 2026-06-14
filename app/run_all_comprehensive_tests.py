# run_all_comprehensive_tests.py
# Master Test Runner - Executes both Appium (Mobile) and Selenium (Web) tests
# Generates comprehensive Excel reports for both and a combined summary

import subprocess
import sys
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from appium_tests.test_cases_100 import TEST_CASES_APPIUM
from automation_tests.test_cases_selenium_100 import TEST_CASES_SELENIUM

class ComprehensiveTestExecutor:
    def __init__(self):
        self.appium_report = "appium_tests/Appium_Test_Report.xlsx"
        self.selenium_report = "automation_tests/Selenium_Test_Report.xlsx"
        self.combined_report = "Test_Execution_Report_Combined.xlsx"
        
    def generate_appium_report_template(self):
        """Generate Appium Excel report template"""
        print("\n" + "="*80)
        print("GENERATING APPIUM TEST REPORT TEMPLATE")
        print("="*80)
        
        try:
            from appium_tests.generate_appium_excel_report import AppiumExcelReportGenerator
            generator = AppiumExcelReportGenerator(self.appium_report)
            generator.generate_report()
            return True
        except Exception as e:
            print(f"✗ Error generating Appium report: {str(e)}")
            return False
    
    def generate_selenium_report_template(self):
        """Generate Selenium Excel report template"""
        print("\n" + "="*80)
        print("GENERATING SELENIUM TEST REPORT TEMPLATE")
        print("="*80)
        
        try:
            from automation_tests.generate_selenium_excel_report import SeleniumExcelReportGenerator
            generator = SeleniumExcelReportGenerator(self.selenium_report)
            generator.generate_report()
            return True
        except Exception as e:
            print(f"✗ Error generating Selenium report: {str(e)}")
            return False
    
    def run_appium_tests(self):
        """Run Appium tests"""
        print("\n" + "="*80)
        print("EXECUTING APPIUM MOBILE TESTS")
        print("="*80)
        
        try:
            from appium_tests.appium_test_runner import AppiumTestRunner
            runner = AppiumTestRunner()
            # Note: In production, this would actually connect to Appium server
            print("✓ Appium test runner initialized")
            print("Note: Connect to Appium server for live testing")
            return True
        except Exception as e:
            print(f"✗ Error running Appium tests: {str(e)}")
            return False
    
    def run_selenium_tests(self):
        """Run Selenium tests"""
        print("\n" + "="*80)
        print("EXECUTING SELENIUM WEB TESTS")
        print("="*80)
        
        try:
            from automation_tests.selenium_test_runner import SeleniumTestRunner
            runner = SeleniumTestRunner()
            # Note: In production, this would actually run against the web application
            print("✓ Selenium test runner initialized")
            print("Note: Ensure web server is running on http://localhost:3000")
            return True
        except Exception as e:
            print(f"✗ Error running Selenium tests: {str(e)}")
            return False
    
    def create_combined_summary_report(self):
        """Create a combined summary report for both testing types"""
        print("\n" + "="*80)
        print("CREATING COMBINED TEST EXECUTION REPORT")
        print("="*80)
        
        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            # Summary Sheet
            self._create_summary_sheet(workbook)
            
            # Appium Test Cases Sheet
            self._create_appium_sheet(workbook)
            
            # Selenium Test Cases Sheet
            self._create_selenium_sheet(workbook)
            
            # Execution Guidelines Sheet
            self._create_guidelines_sheet(workbook)
            
            workbook.save(self.combined_report)
            print(f"✓ Combined report generated: {self.combined_report}")
            return True
        except Exception as e:
            print(f"✗ Error creating combined report: {str(e)}")
            return False
    
    def _create_summary_sheet(self, workbook):
        """Create executive summary sheet"""
        sheet = workbook.create_sheet("Executive Summary", 0)
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = "PATHOAI - END-TO-END TEST EXECUTION REPORT"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Report Info
        sheet['A3'] = "Report Generated:"
        sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet['A4'] = "Application:"
        sheet['B4'] = "PathoAI Clinical Suite"
        sheet['A5'] = "Test Scope:"
        sheet['B5'] = "End-to-End Testing (Mobile + Web)"
        sheet['A6'] = "Testing Framework:"
        sheet['B6'] = "Appium (Mobile) & Selenium (Web)"
        
        # Test Statistics
        sheet['A8'] = "TEST STATISTICS"
        sheet['A8'].font = Font(bold=True, size=12)
        
        sheet['A9'] = "Testing Type"
        sheet['B9'] = "Total Tests"
        sheet['C9'] = "Status"
        sheet['D9'] = "Notes"
        
        for col in ['A', 'B', 'C', 'D']:
            sheet[f'{col}9'].fill = header_fill
            sheet[f'{col}9'].font = header_font
            sheet[f'{col}9'].border = border
        
        # Appium Stats
        sheet['A10'] = "Mobile (Appium)"
        sheet['B10'] = len(TEST_CASES_APPIUM)
        sheet['C10'] = "PENDING"
        sheet['D10'] = "Android native app automation"
        
        # Selenium Stats
        sheet['A11'] = "Web (Selenium)"
        sheet['B11'] = len(TEST_CASES_SELENIUM)
        sheet['C11'] = "PENDING"
        sheet['D11'] = "Web browser automation"
        
        # Combined Total
        sheet['A12'] = "TOTAL"
        sheet['B12'] = len(TEST_CASES_APPIUM) + len(TEST_CASES_SELENIUM)
        sheet['C12'] = "COMPREHENSIVE"
        sheet['D12'] = "Complete end-to-end coverage"
        
        sheet['A12'].font = Font(bold=True)
        sheet['B12'].font = Font(bold=True)
        
        # Categories Summary
        sheet['A14'] = "APPIUM TEST CATEGORIES"
        sheet['A14'].font = Font(bold=True, size=11)
        
        appium_cats = {}
        for test in TEST_CASES_APPIUM:
            cat = test['category']
            appium_cats[cat] = appium_cats.get(cat, 0) + 1
        
        row = 15
        for cat, count in sorted(appium_cats.items()):
            sheet[f'A{row}'] = cat
            sheet[f'B{row}'] = count
            row += 1
        
        # Selenium Categories
        row += 1
        sheet[f'A{row}'] = "SELENIUM TEST CATEGORIES"
        sheet[f'A{row}'].font = Font(bold=True, size=11)
        
        selenium_cats = {}
        for test in TEST_CASES_SELENIUM:
            cat = test['category']
            selenium_cats[cat] = selenium_cats.get(cat, 0) + 1
        
        row += 1
        for cat, count in sorted(selenium_cats.items()):
            sheet[f'A{row}'] = cat
            sheet[f'B{row}'] = count
            row += 1
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 30
    
    def _create_appium_sheet(self, workbook):
        """Create Appium test cases sheet"""
        sheet = workbook.create_sheet("Appium - Mobile Tests")
        self._populate_test_sheet(sheet, TEST_CASES_APPIUM)
    
    def _create_selenium_sheet(self, workbook):
        """Create Selenium test cases sheet"""
        sheet = workbook.create_sheet("Selenium - Web Tests")
        self._populate_test_sheet(sheet, TEST_CASES_SELENIUM)
    
    def _populate_test_sheet(self, sheet, test_cases):
        """Populate test sheet with test cases"""
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        critical_font = Font(color="FFFFFF", bold=True)
        high_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        medium_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        low_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Headers
        headers = ["Test ID", "Category", "Feature", "Priority", "Description", "Steps", 
                  "Expected Result", "Status", "Result", "Execution Date", "Duration (s)", "Comments"]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Data rows
        for row_num, test_case in enumerate(test_cases, 2):
            sheet.cell(row=row_num, column=1).value = test_case['id']
            sheet.cell(row=row_num, column=2).value = test_case['category']
            sheet.cell(row=row_num, column=3).value = test_case['feature']
            sheet.cell(row=row_num, column=4).value = test_case['priority']
            sheet.cell(row=row_num, column=5).value = test_case['description']
            sheet.cell(row=row_num, column=6).value = test_case['steps']
            sheet.cell(row=row_num, column=7).value = test_case['expected']
            sheet.cell(row=row_num, column=8).value = test_case['status']
            sheet.cell(row=row_num, column=9).value = ""
            sheet.cell(row=row_num, column=10).value = ""
            sheet.cell(row=row_num, column=11).value = ""
            sheet.cell(row=row_num, column=12).value = ""
            
            # Apply borders and alignment
            for col in range(1, 13):
                cell = sheet.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Apply priority coloring
            priority_cell = sheet.cell(row=row_num, column=4)
            if priority_cell.value == "CRITICAL":
                priority_cell.fill = critical_fill
                priority_cell.font = critical_font
            elif priority_cell.value == "HIGH":
                priority_cell.fill = high_fill
            elif priority_cell.value == "MEDIUM":
                priority_cell.fill = medium_fill
            elif priority_cell.value == "LOW":
                priority_cell.fill = low_fill
            
            # Apply status coloring
            status_cell = sheet.cell(row=row_num, column=8)
            if status_cell.value == "PENDING":
                status_cell.fill = pending_fill
        
        # Set column widths
        column_widths = {'A': 10, 'B': 18, 'C': 15, 'D': 12, 'E': 25, 'F': 25, 
                        'G': 25, 'H': 12, 'I': 12, 'J': 15, 'K': 12, 'L': 15}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
    
    def _create_guidelines_sheet(self, workbook):
        """Create execution guidelines sheet"""
        sheet = workbook.create_sheet("Execution Guidelines")
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        sheet['A1'] = "TEST EXECUTION GUIDELINES"
        sheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        sheet['A1'].fill = header_fill
        sheet.merge_cells('A1:B1')
        
        # Appium Setup
        sheet['A3'] = "APPIUM TEST SETUP (MOBILE)"
        sheet['A3'].font = Font(bold=True, size=12)
        
        appium_steps = [
            "1. Install Android SDK and Appium Server",
            "2. Connect Android device or start emulator",
            "3. Start Appium server: appium",
            "4. Update app_path and device_id in appium_test_runner.py",
            "5. Run: python appium_tests/appium_test_runner.py",
            "6. Review generated Excel report"
        ]
        
        for i, step in enumerate(appium_steps, 4):
            sheet[f'A{i}'] = step
        
        # Selenium Setup
        row = 4 + len(appium_steps) + 2
        sheet[f'A{row}'] = "SELENIUM TEST SETUP (WEB)"
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        
        selenium_steps = [
            "1. Install ChromeDriver or compatible WebDriver",
            "2. Update BASE_URL in selenium_test_runner.py",
            "3. Start web application server on http://localhost:3000",
            "4. Run: python automation_tests/selenium_test_runner.py",
            "5. Tests will execute in Chrome browser",
            "6. Review generated Excel report"
        ]
        
        for i, step in enumerate(selenium_steps, row + 1):
            sheet[f'A{i}'] = step
        
        # Test Metrics
        row = 4 + len(appium_steps) + 2 + len(selenium_steps) + 2
        sheet[f'A{row}'] = "TEST METRICS & REPORTING"
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        
        metrics = [
            f"• Total Appium Test Cases: {len(TEST_CASES_APPIUM)}",
            f"• Total Selenium Test Cases: {len(TEST_CASES_SELENIUM)}",
            f"• Combined Test Cases: {len(TEST_CASES_APPIUM) + len(TEST_CASES_SELENIUM)}",
            "• Test Priority Levels: CRITICAL, HIGH, MEDIUM, LOW",
            "• Test Categories: Splash, Login, Registration, Dashboard, Analysis, Patient Mgmt, Settings",
            "• Report Format: Excel with multiple sheets",
            "• Execution Time: Tracked per test",
            "• Pass/Fail Status: Automatically marked"
        ]
        
        for i, metric in enumerate(metrics, row + 1):
            sheet[f'A{i}'] = metric
        
        # Best Practices
        row = 4 + len(appium_steps) + 2 + len(selenium_steps) + 2 + len(metrics) + 2
        sheet[f'A{row}'] = "BEST PRACTICES"
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        
        practices = [
            "1. Run Appium tests on actual device when possible",
            "2. Use explicit waits instead of implicit waits",
            "3. Take screenshots on test failures",
            "4. Keep test data separated from test logic",
            "5. Use descriptive element locators",
            "6. Implement proper error handling",
            "7. Document test cases and expected outcomes",
            "8. Run regression tests after code changes"
        ]
        
        for i, practice in enumerate(practices, row + 1):
            sheet[f'A{i}'] = practice
        
        sheet.column_dimensions['A'].width = 70
    
    def print_summary(self):
        """Print execution summary"""
        print("\n" + "="*80)
        print("TEST EXECUTION COMPLETE")
        print("="*80)
        print(f"\n✓ Appium Test Report: {self.appium_report}")
        print(f"  - Total Test Cases: {len(TEST_CASES_APPIUM)}")
        print(f"  - Categories: {', '.join(set(t['category'] for t in TEST_CASES_APPIUM))}")
        
        print(f"\n✓ Selenium Test Report: {self.selenium_report}")
        print(f"  - Total Test Cases: {len(TEST_CASES_SELENIUM)}")
        print(f"  - Categories: {', '.join(set(t['category'] for t in TEST_CASES_SELENIUM))}")
        
        print(f"\n✓ Combined Report: {self.combined_report}")
        print(f"  - Total Test Cases: {len(TEST_CASES_APPIUM) + len(TEST_CASES_SELENIUM)}")
        print(f"  - Includes: Executive Summary, Guidelines, Test Cases")
        
        print("\n" + "="*80)
        print("NEXT STEPS:")
        print("="*80)
        print("1. Connect Appium server and run Appium tests")
        print("2. Start web application and run Selenium tests")
        print("3. Execute tests and review Excel reports")
        print("4. Update test status (PASS/FAIL) in Excel sheets")
        print("5. Generate final test reports for stakeholders")
        print("="*80 + "\n")
    
    def execute_all(self):
        """Execute all test setup"""
        print("\n" + "#"*80)
        print("# PATHOAI END-TO-END TEST EXECUTION SUITE")
        print("#"*80)
        
        # Generate report templates
        self.generate_appium_report_template()
        self.generate_selenium_report_template()
        
        # Create combined report
        self.create_combined_summary_report()
        
        # Print summary
        self.print_summary()

if __name__ == "__main__":
    executor = ComprehensiveTestExecutor()
    executor.execute_all()
