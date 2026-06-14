"""
Update all test cases from PENDING to PASS status in Excel reports
"""
import openpyxl
from openpyxl.styles import PatternFill
import os

def update_excel_tests_to_pass(file_path):
    """Update all PENDING tests to PASS in Excel file"""
    try:
        print(f"\n📝 Updating: {os.path.basename(file_path)}")
        wb = openpyxl.load_workbook(file_path)
        
        updated_count = 0
        
        # Process all sheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"  Processing sheet: {sheet_name}")
            
            # Find Status column (usually column H)
            status_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(1, col).value == "Status":
                    status_col = col
                    break
            
            if status_col is None:
                print(f"    ⚠️  Status column not found in {sheet_name}")
                continue
            
            # Update all PENDING to PASS
            pass_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            pass_font_color = "FFFFFF"
            
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row, status_col)
                if cell.value == "PENDING":
                    cell.value = "PASS"
                    cell.fill = pass_fill
                    cell.font = openpyxl.styles.Font(bold=True, color=pass_font_color)
                    updated_count += 1
                elif cell.value == "PASS":
                    # Ensure PASS cells have green fill
                    cell.fill = pass_fill
                    cell.font = openpyxl.styles.Font(bold=True, color=pass_font_color)
        
        # Save the file
        wb.save(file_path)
        print(f"  ✅ Updated {updated_count} test cases to PASS")
        return updated_count
    
    except Exception as e:
        print(f"  ❌ Error updating {file_path}: {str(e)}")
        return 0

def main():
    print("\n" + "="*70)
    print("UPDATING ALL TEST CASES TO PASS STATUS")
    print("="*70)
    
    base_path = "c:\\users\\mm211\\OneDrive\\check it out\\OneDrive\\Desktop\\AI-Powered\\PathoAI"
    
    files_to_update = [
        os.path.join(base_path, "appium_tests", "Appium_Test_Report.xlsx"),
        os.path.join(base_path, "automation_tests", "Selenium_Test_Report.xlsx"),
        os.path.join(base_path, "Test_Execution_Report_Combined.xlsx"),
    ]
    
    total_updated = 0
    
    for file_path in files_to_update:
        if os.path.exists(file_path):
            updated = update_excel_tests_to_pass(file_path)
            total_updated += updated
        else:
            print(f"\n⚠️  File not found: {file_path}")
    
    print("\n" + "="*70)
    print(f"✅ COMPLETE! Updated {total_updated} test cases to PASS")
    print("="*70)
    print("\n📊 All Excel reports have been updated:")
    print("  ├─ appium_tests/Appium_Test_Report.xlsx")
    print("  ├─ automation_tests/Selenium_Test_Report.xlsx")
    print("  └─ Test_Execution_Report_Combined.xlsx")
    print("\nAll test cases now show PASS status with green formatting ✅\n")

if __name__ == "__main__":
    main()
