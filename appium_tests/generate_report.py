# generate_report.py
# Uses openpyxl to output a premium, beautifully styled Excel spreadsheet containing all test cases.

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(test_cases, output_path="PathoAI_Appium_Test_Report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Execution Report"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # 1. Color Palette Definitions (HSL Tailored/Premium Slate Theme)
    COLOR_PRIMARY_DARK = "0F172A"   # Very dark blue-slate (Title block)
    COLOR_SECONDARY_DARK = "1E293B" # Dark blue-slate (Table headers)
    COLOR_ZEBRA_LIGHT = "F8FAFC"    # Extremely light blue-slate for zebra stripes
    COLOR_WHITE = "FFFFFF"
    COLOR_BORDER = "CBD5E1"         # Slate-200 border color
    
    # Status Colors
    COLOR_PASS_BG = "DCFCE7"        # Soft green
    COLOR_PASS_FG = "15803D"        # Dark green text
    COLOR_FAIL_BG = "FEE2E2"        # Soft red
    COLOR_FAIL_FG = "B91C1C"        # Dark red text
    COLOR_PENDING_BG = "F1F5F9"     # Soft gray
    COLOR_PENDING_FG = "475569"     # Slate gray text

    # Styles
    font_family = "Segoe UI"
    font_title = Font(name=font_family, size=16, bold=True, color=COLOR_WHITE)
    font_subtitle = Font(name=font_family, size=10, italic=True, color="94A3B8")
    font_stats_lbl = Font(name=font_family, size=9, bold=True, color="64748B")
    font_stats_val = Font(name=font_family, size=11, bold=True, color=COLOR_PRIMARY_DARK)
    
    font_header = Font(name=font_family, size=11, bold=True, color=COLOR_WHITE)
    font_body = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    
    fill_title = PatternFill(start_color=COLOR_PRIMARY_DARK, end_color=COLOR_PRIMARY_DARK, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_SECONDARY_DARK, end_color=COLOR_SECONDARY_DARK, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA_LIGHT, end_color=COLOR_ZEBRA_LIGHT, fill_type="solid")
    fill_white = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    
    # Status fills and fonts
    fill_pass = PatternFill(start_color=COLOR_PASS_BG, end_color=COLOR_PASS_BG, fill_type="solid")
    font_pass = Font(name=font_family, size=10, bold=True, color=COLOR_PASS_FG)
    
    fill_fail = PatternFill(start_color=COLOR_FAIL_BG, end_color=COLOR_FAIL_BG, fill_type="solid")
    font_fail = Font(name=font_family, size=10, bold=True, color=COLOR_FAIL_FG)
    
    fill_pending = PatternFill(start_color=COLOR_PENDING_BG, end_color=COLOR_PENDING_BG, fill_type="solid")
    font_pending = Font(name=font_family, size=10, italic=True, color=COLOR_PENDING_FG)

    # Borders
    thin_side = Side(style='thin', color=COLOR_BORDER)
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_bottom_side = Side(style='medium', color="475569")
    
    # 2. Add Title Block Banner (Rows 1-4)
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = "BIOPATH AI — MOBILE END-TO-END AUTOMATION REPORT"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fill remaining merged cells to apply background color properly
    for row in range(1, 3):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = fill_title

    # 3. Execution Stats Panel (Row 4)
    total_tests = len(test_cases)
    passed_tests = sum(1 for tc in test_cases if tc["status"] == "PASS")
    failed_tests = sum(1 for tc in test_cases if tc["status"] == "FAIL")
    pending_tests = sum(1 for tc in test_cases if tc["status"] in ["PENDING", "BLOCKED"])
    pass_rate = (passed_tests / total_tests) * 100 if total_tests > 0 else 0

    stats = [
        ("TOTAL CASES", total_tests, "A4:B4"),
        ("PASSED", passed_tests, "C4:D4"),
        ("FAILED", failed_tests, "E4:F4"),
        ("PASS RATE", f"{pass_rate:.1f}%", "G4:H4")
    ]

    for label, val, rng in stats:
        start_cell_ref = rng.split(":")[0]
        ws.merge_cells(rng)
        cell = ws[start_cell_ref]
        cell.value = f"{label}: {val}"
        cell.font = Font(name=font_family, size=10, bold=True, color=COLOR_PRIMARY_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Apply light gray background & border to stats row
        for r in range(4, 5):
            parts = rng.split(":")
            start_col = ord(parts[0][0]) - ord('A') + 1
            end_col = ord(parts[1][0]) - ord('A') + 1
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                ws.cell(row=r, column=c).border = Border(bottom=thick_bottom_side, top=thin_side, left=thin_side, right=thin_side)

    # Row spacing
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 25

    # 4. Table Headers (Row 6)
    headers = [
        "Test ID",
        "Category",
        "Feature",
        "Description",
        "Execution Steps",
        "Expected Outcome",
        "Status",
        "Execution Comments"
    ]
    
    header_row = 6
    ws.row_dimensions[header_row].height = 28
    
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if idx in [1, 7] else "left", vertical="center")
        cell.border = border_all

    # 5. Populate Test Case Rows
    current_row = 7
    for idx, tc in enumerate(test_cases):
        ws.row_dimensions[current_row].height = 36  # Generous height for readability
        
        # Zebra Striping logic
        row_fill = fill_zebra if idx % 2 == 1 else fill_white
        
        # Values
        row_values = [
            tc["id"],
            tc["category"],
            tc["feature"],
            tc["description"],
            tc["steps"],
            tc["expected"],
            tc["status"],
            tc["comments"]
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = font_body
            cell.fill = row_fill
            cell.border = border_all
            
            # Alignments
            if col_idx == 1:
                # Test ID centered and bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font_bold
            elif col_idx == 7:
                # Status centered with custom fills/fonts
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "PASS":
                    cell.fill = fill_pass
                    cell.font = font_pass
                elif val == "FAIL":
                    cell.fill = fill_fail
                    cell.font = font_fail
                else:
                    cell.fill = fill_pending
                    cell.font = font_pending
            else:
                # Text fields left aligned with wrap
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        current_row += 1

    # 6. Auto-fit column widths dynamically with padding
    col_widths = {
        1: 10,  # ID
        2: 20,  # Category
        3: 20,  # Feature
        4: 35,  # Description
        5: 45,  # Steps
        6: 40,  # Expected
        7: 12,  # Status
        8: 30   # Comments
    }
    
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Save to disk
    wb.save(output_path)
    print(f"Excel report successfully generated at: {output_path}")

if __name__ == "__main__":
    from test_cases import TEST_CASES
    # Simulating standard generation with raw pending cases
    generate_excel_report(TEST_CASES, "PathoAI_Appium_Test_Report.xlsx")
