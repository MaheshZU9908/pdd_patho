'''\n# generate_report.py - Create an Excel report summarizing Appium & Selenium test results\n\nimport os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
\n\ndef write_report(results, output_file=None):\n    """Write test results to an Excel workbook.
\n    Args:\n        results (list of dict): Each dict should contain keys: 'id', 'description', 'type', 'status'.\n        output_file (str or Path, optional): Destination path for the Excel file.\n            If None, the file is written to 'test_report.xlsx' in the current working directory.\n    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"
\n    # Header row
    headers = ["Test ID", "Description", "Type", "Status"]
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="FF4F81BD")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
\n    # Data rows
    for row_idx, item in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("id"))
        ws.cell(row=row_idx, column=2, value=item.get("description"))
        ws.cell(row=row_idx, column=3, value=item.get("type"))
        status_cell = ws.cell(row=row_idx, column=4, value=item.get("status"))
        # Color‑code status
        if item.get("status") == "Pass":
            status_cell.font = Font(color="FF006100")  # dark green
        else:
            status_cell.font = Font(color="FF9C0006")  # dark red
\n    # Auto‑size columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = length + 2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
\n    # Determine output path
    if output_file is None:
        output_path = Path.cwd() / "test_report.xlsx"
    else:
        output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Report written to {output_path}")
    return output_path
\n\nif __name__ == "__main__":\n    # Demo – generate an empty report if run directly
    demo_results = [\n        {"id": "M001", "description": "Demo mobile test", "type": "mobile", "status": "Pass"},\n        {"id": "W001", "description": "Demo web test", "type": "web", "status": "Fail"}\n    ]\n    write_report(demo_results)\n'''
