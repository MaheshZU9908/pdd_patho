# generate_selenium_excel_report.py
# Generate comprehensive Excel report for Selenium test cases

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from automation_tests.test_cases_selenium_100 import TEST_CASES_SELENIUM
import os

class SeleniumExcelReportGenerator:
    def __init__(self, output_path="automation_tests/Selenium_Test_Report.xlsx"):
        self.output_path = output_path
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Selenium Tests"
        
    def setup_styles(self):
        """Setup cell styles"""
        self.header_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        self.critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.critical_font = Font(color="FFFFFF", bold=True)
        
        self.high_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.medium_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.low_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
    def add_headers(self):
        """Add header row"""
        headers = [
            "Test ID",
            "Category",
            "Feature",
            "Priority",
            "Description",
            "Steps",
            "Expected Result",
            "Status",
            "Execution Date",
            "Result",
            "Comments",
            "Duration (sec)",
            "Error/Issue"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
    
    def set_column_widths(self):
        """Set column widths"""
        column_widths = {
            'A': 12,  # Test ID
            'B': 18,  # Category
            'C': 18,  # Feature
            'D': 12,  # Priority
            'E': 25,  # Description
            'F': 30,  # Steps
            'G': 30,  # Expected Result
            'H': 12,  # Status
            'I': 15,  # Execution Date
            'J': 12,  # Result
            'K': 20,  # Comments
            'L': 12,  # Duration
            'M': 20   # Error/Issue
        }
        
        for col_letter, width in column_widths.items():
            self.sheet.column_dimensions[col_letter].width = width
    
    def add_test_data(self):
        """Add test data rows"""
        for row_num, test_case in enumerate(TEST_CASES_SELENIUM, 2):
            self.sheet.cell(row=row_num, column=1).value = test_case['id']
            self.sheet.cell(row=row_num, column=2).value = test_case['category']
            self.sheet.cell(row=row_num, column=3).value = test_case['feature']
            self.sheet.cell(row=row_num, column=4).value = test_case['priority']
            self.sheet.cell(row=row_num, column=5).value = test_case['description']
            self.sheet.cell(row=row_num, column=6).value = test_case['steps']
            self.sheet.cell(row=row_num, column=7).value = test_case['expected']
            self.sheet.cell(row=row_num, column=8).value = test_case['status']
            self.sheet.cell(row=row_num, column=9).value = ""  # Execution Date
            self.sheet.cell(row=row_num, column=10).value = "PENDING"  # Result
            self.sheet.cell(row=row_num, column=11).value = ""  # Comments
            self.sheet.cell(row=row_num, column=12).value = ""  # Duration
            self.sheet.cell(row=row_num, column=13).value = ""  # Error/Issue
            
            # Apply borders
            for col in range(1, 14):
                self.sheet.cell(row=row_num, column=col).border = self.border
                self.sheet.cell(row=row_num, column=col).alignment = self.left_align
    
    def apply_conditional_formatting(self):
        """Apply priority-based coloring"""
        for row_num in range(2, len(TEST_CASES_SELENIUM) + 2):
            priority_cell = self.sheet.cell(row=row_num, column=4)
            priority = priority_cell.value
            
            if priority == "CRITICAL":
                priority_cell.fill = self.critical_fill
                priority_cell.font = self.critical_font
            elif priority == "HIGH":
                priority_cell.fill = self.high_fill
            elif priority == "MEDIUM":
                priority_cell.fill = self.medium_fill
            elif priority == "LOW":
                priority_cell.fill = self.low_fill
            
            # Status coloring
            status_cell = self.sheet.cell(row=row_num, column=8)
            if status_cell.value == "PENDING":
                status_cell.fill = self.pending_fill
            elif status_cell.value == "PASS":
                status_cell.fill = self.pass_fill
            elif status_cell.value == "FAIL":
                status_cell.fill = self.fail_fill
    
    def add_summary_sheet(self):
        """Add summary statistics sheet"""
        summary_sheet = self.workbook.create_sheet("Summary")
        
        # Title
        summary_sheet['A1'] = "SELENIUM TEST EXECUTION SUMMARY"
        summary_sheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        summary_sheet['A1'].fill = self.header_fill
        summary_sheet.merge_cells('A1:B1')
        
        # Test counts by category
        summary_sheet['A3'] = "TEST STATISTICS"
        summary_sheet['A3'].font = Font(bold=True, size=12)
        
        categories = {}
        for test in TEST_CASES_SELENIUM:
            cat = test['category']
            categories[cat] = categories.get(cat, 0) + 1
        
        row = 4
        summary_sheet['A4'] = "Category"
        summary_sheet['B4'] = "Count"
        summary_sheet['A4'].fill = self.header_fill
        summary_sheet['B4'].fill = self.header_fill
        summary_sheet['A4'].font = self.header_font
        summary_sheet['B4'].font = self.header_font
        
        for category, count in sorted(categories.items()):
            row += 1
            summary_sheet[f'A{row}'] = category
            summary_sheet[f'B{row}'] = count
        
        # Priority breakdown
        summary_sheet['D3'] = "PRIORITY BREAKDOWN"
        summary_sheet['D3'].font = Font(bold=True, size=12)
        
        priorities = {}
        for test in TEST_CASES_SELENIUM:
            pri = test['priority']
            priorities[pri] = priorities.get(pri, 0) + 1
        
        summary_sheet['D4'] = "Priority"
        summary_sheet['E4'] = "Count"
        summary_sheet['D4'].fill = self.header_fill
        summary_sheet['E4'].fill = self.header_fill
        summary_sheet['D4'].font = self.header_font
        summary_sheet['E4'].font = self.header_font
        
        row = 4
        for priority in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
            row += 1
            summary_sheet[f'D{row}'] = priority
            summary_sheet[f'E{row}'] = priorities.get(priority, 0)
        
        # Report details
        summary_sheet['A11'] = "Report Details"
        summary_sheet['A11'].font = Font(bold=True, size=12)
        
        summary_sheet['A12'] = "Generated Date:"
        summary_sheet['B12'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        summary_sheet['A13'] = "Total Test Cases:"
        summary_sheet['B13'] = len(TEST_CASES_SELENIUM)
        
        summary_sheet['A14'] = "Application:"
        summary_sheet['B14'] = "PathoAI Web Application"
        
        summary_sheet['A15'] = "Test Type:"
        summary_sheet['B15'] = "Selenium Web Automation Testing"
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['D'].width = 25
        summary_sheet.column_dimensions['E'].width = 15
    
    def generate_report(self):
        """Generate complete report"""
        self.setup_styles()
        self.add_headers()
        self.set_column_widths()
        self.add_test_data()
        self.apply_conditional_formatting()
        self.add_summary_sheet()
        
        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(self.output_path) if os.path.dirname(self.output_path) else ".", exist_ok=True)
        
        # Save workbook
        self.workbook.save(self.output_path)
        print(f"✓ Selenium test report generated: {self.output_path}")
        return self.output_path

if __name__ == "__main__":
    generator = SeleniumExcelReportGenerator()
    generator.generate_report()
